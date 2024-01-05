from django.contrib import admin
from .models import Fornecedor
from openpyxl import Workbook
from django.http import HttpResponse

def exportar_para_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=fornecedores.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fornecedores'

    # Escrevendo os cabeçalhos
    colunas = ['CNPJ', 'Categoria', 'Empresa', 'Endereço', 'CEP', 'Email', 'Contato', 'Tempo Atividade']
    for col_num, column_title in enumerate(colunas, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Escrevendo os dados
    for row_num, fornecedor in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=fornecedor.cnpj)
        ws.cell(row=row_num, column=2, value=fornecedor.categoria)
        ws.cell(row=row_num, column=3, value=fornecedor.empresa)
        ws.cell(row=row_num, column=4, value=fornecedor.endereco)
        ws.cell(row=row_num, column=5, value=fornecedor.cep)
        ws.cell(row=row_num, column=6, value=fornecedor.email)
        ws.cell(row=row_num, column=7, value=fornecedor.contato)
        ws.cell(row=row_num, column=8, value=fornecedor.tempo_atividade)

    wb.save(response)
    return response

exportar_para_excel.short_description = 'Exportar selecionados para Excel'



class FornecedorAdmin(admin.ModelAdmin):
    list_display = ('empresa', 'cnpj', 'email', 'contato', 'tempo_atividade')  # Campos a serem exibidos
    search_fields = ['empresa', 'cnpj']  # Campos de pesquisa
    actions = [exportar_para_excel]

# Registra o modelo Fornecedor para que ele apareça na interface admin
admin.site.register(Fornecedor, FornecedorAdmin)
